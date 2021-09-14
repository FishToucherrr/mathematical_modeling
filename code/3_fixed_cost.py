import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pulp
import openpyxl


data    =   pd.read_excel("附件1 近5年402家供应商的相关数据.xlsx",
                sheet_name='供应商的供货量（m³）')
data_1  =   pd.read_excel("附件1 近5年402家供应商的相关数据.xlsx",
                sheet_name='企业的订货量（m³）')
data_2  =   pd.read_excel("附件2 近5年8家转运商的相关数据.xlsx",
                sheet_name = '运输损耗率（%）')

data    =   data.iloc[0:,1:]
data_1  =   data_1.iloc[0:,1:]
data_2  =   data_2.iloc[0:,1:]

data=np.array(data)
data_1=np.array(data_1)
data_2=np.array(data_2)

ratio = 1                       #原料转换比
data_ref=np.zeros((402,240))                  #将ABC的供货量转化为产品的生产量

production = np.zeros(240)      #每周总原料量
production_deal = np.zeros(240) #每周去掉特定生产商的原料量

stock = 0                       #库存
difference = 0                  #差值 中间变量
result = np.zeros(402)

for i in range(402):
    if data[i][0]=='A':
        ratio = 0.6
    elif data[i][0]=='B':
        ratio = 0.66
    else:
        ratio = 0.72
    
    for j in range(240):
        data_ref[i][j] = data[i][j+1]/ratio

for i in range(240):
    for j in range(402):
        production[i]+= data_ref[j][i]

for j in range(402):
    stock = 0
    difference = 0
    for i in range(240):
        production_deal[i]=production[i]-data_ref[j][i]
    for i in range(240):
        if stock+production_deal[i]-28200 > 0:
            stock=stock+production_deal[i]-28200
            difference = 0
        else:
            difference = 28200 - stock - production_deal[i]
            stock = 0
        result[j]+=difference**2
    result[j]=(result[j]/240)**0.5

num = np.arange(402)

num_result = zip(result,num)
list_result = list(num_result)


list_result=sorted(list_result,key=(lambda x:x[0]),reverse=True)


first_50 = []

for i in range(50):
    first_50.append(list_result[i][1])               #此处供应商序号从0开始，非从1开始

production_for_window=np.zeros(240)

count=0
data_for_window=np.zeros((50,240))

for i in range(402):
    if i in first_50:
        data_for_window[count]=data_ref[i]
        count+=1
        for j in range(240):
            production_for_window[j]+=data_ref[i][j]


satisfy=0

window_width=24     #滑动窗口宽度
start_week_set=[]      #满足条件的窗口 起始月份集合


for i in range(240-window_width):
    index=i
    stock=0
    satisfy=1
    while index < i+window_width :
        if stock + production[index] > 2*28200:
            stock=stock + production[index]-28200
            index+=1
        else:
            satisfy=0
            break
    if satisfy==1:
        start_week_set.append(i)

#print(start_week_set)
#print(production_for_window[82])
'''
plt.title('supply')  # 折线图标题
plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示汉字
plt.xlabel('week')  # x轴标题
plt.ylabel('supply')  # y轴标题

'''
# sol_list=[]
'''
for start_week in start_week_set:

    sum_matrix = np.zeros((50,24))

    for i in range(50):
        sum_matrix[i][0]=data_for_window[i][start_week]

    for i in range(50):
        for j in range(start_week+1,start_week+24):
            sum_matrix[i][j-start_week]=sum_matrix[i][j-1-start_week]+data_for_window[i][j]

    InvestLP = pulp.LpProblem("problem_for_window", sense=pulp.LpMinimize)

    types=[str(i) for i in range(1,51)]
    status = pulp.LpVariable.dicts("supplier",types,cat='Binary')

    InvestLP += pulp.lpSum([(status[i]) for i in types])


    for i in range(24):
        InvestLP += ((pulp.lpSum([(status[types[j]]*sum_matrix[j][i]) for j in range(50)])-i*28200)>=0)

                    
    InvestLP.solve()

    print(InvestLP.name)
    print("Status :", pulp.LpStatus[InvestLP.status])  # 输出求解状态
    for v in InvestLP.variables():
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值
    
    print("Min f(x) =", pulp.value(InvestLP.objective))  # 输出最优解的目标函数值

    #sol_list.append(pulp.value(InvestLP.objective))
    #if pulp.value(InvestLP.objective) == 21:

#print(sol_list)
#print(len(start_week_set))
'''
start_week=start_week_set[2]

sum_matrix = np.zeros((50,24))

for i in range(50):
    sum_matrix[i][0]=data_for_window[i][start_week]

for i in range(50):
    for j in range(start_week+1,start_week+24):
        sum_matrix[i][j-start_week]=sum_matrix[i][j-1-start_week]+data_for_window[i][j]

InvestLP = pulp.LpProblem("problem_for_window", sense=pulp.LpMinimize)

types=['01','02','03','04','05','06','07','08','09']
for i in range(10,51):
    types.append(str(i))


status = pulp.LpVariable.dicts("supplier",types,cat='Binary')

InvestLP += pulp.lpSum([(status[i]) for i in types])


for i in range(24):
    InvestLP += ((pulp.lpSum([(status[types[j]]*sum_matrix[j][i]) for j in range(50)])-i*28200)>=28200)

                
InvestLP.solve()
print(InvestLP.name)
print("Status :", pulp.LpStatus[InvestLP.status])  # 输出求解状态
for v in InvestLP.variables():
    print(v.name, "=", v.varValue)  # 输出每个变量的最优值
print("Min f(x) =", pulp.value(InvestLP.objective))  # 输出最优解的目标函数值

sup_list=[]

count = 0
for v in InvestLP.variables():
    if v.varValue==1:
        sup_list.append(count)
    count+=1


########################################################################################################################################
real_list=[]

for i in sup_list:
    real_list.append(first_50[i])




count=0

least_suppliers=[]

for v in InvestLP.variables():
    if v.varValue==1:
        least_suppliers.append(count)
    count+=1

num_suppliers=len(least_suppliers)

material_type=[]
ratio_type={}
price_type={}

ratio_type['A']=0.6
ratio_type['B']=0.66
ratio_type['C']=0.72
price_type['A']=1.2
price_type['B']=1.1
price_type['C']=1


real_suppliers=[]

for i in least_suppliers:
    real_suppliers.append(first_50[i])


for i in real_suppliers:
    material_type.append(data[i][0])


data_for_supply= np.zeros((num_suppliers,240)) #选定的21家供货商240周的供货情况 未转换为产品量
data_for_order = np.zeros((num_suppliers,240)) #选定的21家供货商240周的订货情况 未转换为产品量

count=0
for i in real_suppliers:
        data_for_order[count]=data_1[i][1:]
        data_for_supply[count]=data[i][1:]
        count+=1

hist_max=np.zeros(num_suppliers) # 选定的供货商历史供货峰值(转化为产品)

data_for_xxx= np.zeros((num_suppliers,240))

count=0
for i in real_suppliers:
    data_for_xxx[count]=data_ref[i]
    count+=1

for i in range(num_suppliers):
        for k in range(240):
            hist_max[i]=max(hist_max[i],data_for_xxx[i][k])

#print(hist_max)

s=np.zeros(24)
for i in range(num_suppliers):
    for j in range(24):
        s[j]+=hist_max[j]



one_list=np.zeros(24)

stock=np.zeros(24)
for i in range(1,24):
    stock[i]=stock[i-1]+s[i-1]-28200
    if stock[i] >=28200:
        one_list[i]=1

#print(real_suppliers)
#print(stock)
#print(one_list)
#print(len(start_week_set))
#print(start_week_set)

avr_matrix=np.zeros((num_suppliers,24))

for i in range(num_suppliers):
    for j in range(24):
        for k in range(5):
            avr_matrix[i][j]+=data_for_xxx[i][k*48+j]
        avr_matrix[i][j]/=5

base_value=np.zeros(num_suppliers) #平均值峰值 系数计算基准

for i in range(num_suppliers):
    for j in range(24):
        base_value[i]=max(base_value[i],avr_matrix[i][j])

coefficient_matrix=np.zeros((num_suppliers,24))

for i in range(num_suppliers):
    for j in range(24):
        coefficient_matrix[i][j]=(avr_matrix[i][j]/base_value[i])*0.88+0.12

Bound=np.zeros((num_suppliers,24))

real_max=np.zeros(num_suppliers) # 选定的供货商历史供货峰值(转化为产品)

data_for_bound= np.zeros((num_suppliers,240))

count=0
for i in real_suppliers:
    data_for_bound[count]=data[i][1:]
    count+=1

for i in range(num_suppliers):
    for k in range(240):
        real_max[i]=max(real_max[i],data_for_bound[i][k])

for i in range(num_suppliers):
    for j in range(24):
        Bound[i][j]=real_max[i]*coefficient_matrix[i][j]
        #Bound[i][j]=real_max[i]


stock_for_cost=np.zeros(24)

status_list=[]

result_2_2=np.zeros((num_suppliers,24))

min_2_2=[]

for index in range(24):
    costLP=pulp.LpProblem("cost_problem_"+str(index),sense=pulp.LpMinimize)

    yield_list=[]

    for i in range(num_suppliers):
        if i < 10:
            yield_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=40000,cat='Integer'))
        else:
            yield_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=40000,cat='Integer'))

    costLP += pulp.lpSum([(yield_list[i]*price_type[material_type[i]])for i in range(num_suppliers)])

    costLP += ((pulp.lpSum([(yield_list[i]*(ratio_type[material_type[i]])**(-1) )for i in range(num_suppliers)])+stock_for_cost[index]-28200) >= 28200)

    for i in range(num_suppliers):
        costLP += (yield_list[i]<=Bound[i][index])

    costLP.solve()
    print(costLP.name)
    print("Status :", pulp.LpStatus[costLP.status])  # 输出求解状态
    count=0
    for v in costLP.variables():
        result_2_2[count][index]=v.varValue
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值
        count+=1

    print("Min f(x) =", pulp.value(costLP.objective))  # 输出最优解的目标函数值

    min_2_2.append(pulp.value(costLP.objective))
    
    status_list.append(pulp.LpStatus[costLP.status])

    sum_prod=0
    count=0
    for v in costLP.variables():
        sum_prod+=v.varValue/ratio_type[material_type[count]]
        count+=1
    if index <23:
        stock_for_cost[index+1]=stock_for_cost[index]+sum_prod-28200

#print(status_list)


#################################################################################################################################################

prob=np.zeros(num_suppliers)
count=0

for i in real_suppliers:
    temp_sup=0
    temp_ord=0
    for j in range(240):
        if data_1[i][j+1]>0 and data[i][j+1]/data_1[i][j+1] >= 0.8 :
            temp_ord+=data_1[i][j+1]
            temp_sup+=data[i][j+1]
    prob[count]=temp_sup/temp_ord
    count+=1

#print(prob)

#for i in range(num_suppliers):
    #result_2_2[i][0]=real_suppliers[i]
    #for j in range(24):
    #    result_2_2[i][j+1]=int(result_2_2[i][j+1]/prob[i])

dealt_shipper=np.zeros((8,24))

avg_list=[1,2,3]

for i in range(8):
    if i in avg_list:
        count=0
        Sum=0
        for k in range(240):
            if data_2[i][k]>0:
                count+=1
                Sum+=data_2[i][k]
        Sum=Sum/count
        for j in range(24):
            dealt_shipper[i][j]=Sum
    else:
        for j in range(24):
            count=0
            for k in range(10):
                if data_2[i][24*k+j]!=0:
                    dealt_shipper[i][j]+=data_2[i][24*k+j]
                    count+=1
            if count==0:
                c=0
                for k in range(240):
                    if data_2[i][k]!=0:
                        c+=1
                        dealt_shipper[i][j]+=data_2[i][k]
                dealt_shipper[i][j]/=c
            else:
                dealt_shipper[i][j]/=count

#print(dealt_shipper)

result_2_3=[]

for index in range(24):

    lossLP=pulp.LpProblem("loss_problem_"+str(index),sense=pulp.LpMinimize)

    dec_list=[]

    par=np.zeros(num_suppliers*8)

    for i in range(num_suppliers*8):
        par[i]=(dealt_shipper[i%8][index]/100)/ratio_type[material_type[int(i/8)]]

    #print(dealt_shipper)
    #print(par)

    for i in range(8*num_suppliers):
        if i < 10:
            dec_list.append(pulp.LpVariable('00'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        elif i <100:
            dec_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        else:
            dec_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=6000,cat='Integer'))
       
    for i in range(num_suppliers):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for j in range(8)])==result_2_2[i][index])

    for j in range(8):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for i in range(num_suppliers)])<=6000)
    
    lossLP += pulp.lpSum([(dec_list[i]*par[i])  for i in range(num_suppliers*8)])

    
    lossLP.solve()
    
    print(lossLP.name)
    print("Status :", pulp.LpStatus[lossLP.status])  # 输出求解状态
    for v in lossLP.variables():
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值

    print("Min f(x) =", pulp.value(lossLP.objective))  # 输出最优解的目标函数值
    
    rlt=np.zeros((num_suppliers,8))

    #print(par)
    count=0
    for v in lossLP.variables():
        rlt[int(count/8)][count%8]=v.varValue
        count+=1
        #print(v.name, "=", v.varValue)  # 输出每个变量的最优值
    
    result_2_3.append(rlt)
    #print(rlt)
    

######################################################################################################################


stock_for_cost=np.zeros(24)

status_list=[]

result_3_1=np.zeros((num_suppliers,24))

for index in range(24):
    costLP=pulp.LpProblem("cost_problem_2_"+str(index),sense=pulp.LpMinimize)

    yield_list=[]

    for i in range(num_suppliers):
        if i < 10:
            yield_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=40000,cat='Integer'))
        else:
            yield_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=40000,cat='Integer'))


    C=20  #转运+仓储的固定成本 参数
    costLP += pulp.lpSum([(yield_list[i]*(price_type[material_type[i]]+C))for i in range(num_suppliers)])

    costLP += ((pulp.lpSum([(yield_list[i]*(ratio_type[material_type[i]])**(-1) )for i in range(num_suppliers)])+stock_for_cost[index]-28200) >= 28200)

    for i in range(num_suppliers):
        costLP += (yield_list[i]<=Bound[i][index])

    costLP.solve()
    print(costLP.name)
    print("Status :", pulp.LpStatus[costLP.status])  # 输出求解状态
    count=0
    for v in costLP.variables():
        result_3_1[count][index]=v.varValue
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值
        count+=1

    print("Min f(x) =", pulp.value(costLP.objective))  # 输出最优解的目标函数值
    
    status_list.append(pulp.LpStatus[costLP.status])

    sum_prod=0
    count=0
    for v in costLP.variables():
        sum_prod+=v.varValue/ratio_type[material_type[count]]
        count+=1
    if index <23:
        stock_for_cost[index+1]=stock_for_cost[index]+sum_prod-28200

print(status_list)

##########################################################################################################################################


#print(dealt_shipper)

result_3_2 = [] 

for index in range(24):

    lossLP=pulp.LpProblem("loss_problem_"+str(index),sense=pulp.LpMinimize)

    dec_list=[]

    par=np.zeros(num_suppliers*8)

    for i in range(num_suppliers*8):
        par[i]=(dealt_shipper[i%8][index]/100)/ratio_type[material_type[int(i/8)]]

    #print(dealt_shipper)
    #print(par)

    for i in range(8*num_suppliers):
        if i < 10:
            dec_list.append(pulp.LpVariable('00'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        elif i <100:
            dec_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        else:
            dec_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=6000,cat='Integer'))
       
    for i in range(num_suppliers):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for j in range(8)])==result_3_1[i][index])

    for j in range(8):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for i in range(num_suppliers)])<=6000)
    
    lossLP += pulp.lpSum([(dec_list[i]*par[i])  for i in range(num_suppliers*8)])

    lossLP.solve()
    print(lossLP.name)
    print("Status :", pulp.LpStatus[lossLP.status])  # 输出求解状态
    for v in lossLP.variables():
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值
    print("Min f(x) =", pulp.value(lossLP.objective))  # 输出最优解的目标函数值
    
    rlt=np.zeros((num_suppliers,8))

    count=0
    for v in lossLP.variables():
        rlt[int(count/8)][count%8]=v.varValue
        count+=1
        #print(v.name, "=", v.varValue)  # 输出每个变量的最优值
    result_3_2.append(rlt)

#print(result_3_2[0])


print(prob)
print(real_list)

display_2_2=np.zeros((num_suppliers,24))
display_3_1=np.zeros((num_suppliers,24))

for i in range(num_suppliers):
    for j in range(24):
        display_2_2[i][j]=int(result_2_2[i][j]/prob[i])

for i in range(num_suppliers):
    for j in range(24):
        display_3_1[i][j]=int(result_3_1[i][j]/prob[i])

c_list=np.zeros(24)

for i in range(num_suppliers):
    if data[real_suppliers[i]][0]=='C':
        for j in range(24):
            c_list[j]+=result_3_1[i][j]

for i in range(24):
    print(c_list[i])

    

workbook=openpyxl.load_workbook("附件A 订购方案数据结果.xlsx")
worksheet=workbook.worksheets[0]

for i in range(num_suppliers):
    for j in range(24):
        x=real_suppliers[i]
        worksheet.cell(7+x,2+j,display_2_2[i][j])

worksheet=workbook.worksheets[1]

for i in range(num_suppliers):
    for j in range(24):
        x=real_suppliers[i]
        worksheet.cell(7+x,2+j,display_3_1[i][j])


workbook.save(filename="附件A 订购方案数据结果.xlsx")

workbook=openpyxl.load_workbook("附件B 转运方案数据结果.xlsx")
worksheet=workbook.worksheets[0]


for i in range(num_suppliers):
    for j in range(24):
        for k in range(8):
            x=real_suppliers[i]
            worksheet.cell(7+x,2+8*j+k,result_2_3[j][i][k])

worksheet=workbook.worksheets[1]


for i in range(num_suppliers):
    for j in range(24):
        for k in range(8):
            x=real_suppliers[i]
            worksheet.cell(7+x,2+8*j+k,result_3_2[j][i][k])


workbook.save(filename="附件B 转运方案数据结果.xlsx")