import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pulp
import openpyxl


data    =   pd.read_excel("附件1 近5年402家供应商的相关数据.xlsx",
                sheet_name='供应商的供货量（m³）')
data_1  =   pd.read_excel("附件1 近5年402家供应商的相关数据.xlsx",
                sheet_name='企业的订货量（m³）')
data_2  =   pd.read_excel("附件2 近5年8家转运商的相关数据.xlsx",
                sheet_name = '运输损耗率（%）')

data    =   data.iloc[0:,1:]
data_1  =   data_1.iloc[0:,1:]
data_2  =   data_2.iloc[0:,1:]

data=np.array(data)
data_1=np.array(data_1)
data_2=np.array(data_2)


hist_max=np.zeros(402) # 选定的供货商历史供货峰值(转化为产品)

for i in range(402):
    for j in range(240):
        hist_max[i]=max(hist_max[i],data[i][j+1])


#print(hist_max)

prob=np.zeros(402)
count=0

for i in range(402):
    temp_sup=0
    temp_ord=0
    for j in range(240):
        if data_1[i][j+1]>0 and data[i][j+1]/data_1[i][j+1] >= 0.8 :
            temp_ord+=data_1[i][j+1]
            temp_sup+=data[i][j+1]
    if temp_ord==0:
        prob[i]=0
    else:
        prob[i]=temp_sup/temp_ord
    count+=1


#print(real_suppliers)
#print(stock)
#print(one_list)
#print(len(start_week_set))
#print(start_week_set)

avr_matrix=np.zeros((402,24))

for i in range(402):
    for j in range(24):
        for k in range(5):
            avr_matrix[i][j]+=data[i][k*48+j+1]
        avr_matrix[i][j]/=5

base_value=np.zeros(402) #平均值峰值 系数计算基准

for i in range(402):
    for j in range(24):
        base_value[i]=max(base_value[i],avr_matrix[i][j])

coefficient_matrix=np.zeros((402,24))

for i in range(402):
    for j in range(24):
        if base_value[i]!=0:
            coefficient_matrix[i][j]=(avr_matrix[i][j]/base_value[i])
        else:
            coefficient_matrix[i][j]=0

Bound=np.zeros((402,24))


for i in range(402):
    for j in range(24):
        Bound[i][j]=int(hist_max[i]*coefficient_matrix[i][j])

ratio={}
ratio['A']=0.60
ratio['B']=0.66
ratio['C']=0.72

abc_Bound=np.zeros((3,24))

for i in range(402):
    if data[i][0] == 'A':
        for j in range(24):
            abc_Bound[0][j]+=Bound[i][j]
    if data[i][0] == 'B':
        for j in range(24):
            abc_Bound[1][j]+=Bound[i][j]
    if data[i][0] == 'C':
        for j in range(24):
            abc_Bound[2][j]+=Bound[i][j]

prod=np.zeros(24)

for i in range(24):
    A_n=0
    B_n=0
    C_n=0
    if abc_Bound[0][i]+abc_Bound[1][i]+abc_Bound[2][i] <=48000:
        A_n=abc_Bound[0][i]
        B_n=abc_Bound[1][i]
        C_n=abc_Bound[2][i]
    elif abc_Bound[0][i]>=48000:
        A_n=48000
        B_n=0
        C_n=0
    elif abc_Bound[0][i]+abc_Bound[1][i]>=48000:
        A_n=abc_Bound[0][i]
        B_n=48000-abc_Bound[0][i]
        C_n=0
    else:
        A_n=abc_Bound[0][i]
        B_n=abc_Bound[0][i]
        C_n=48000-A_n-B_n
    prod[i]=A_n/ratio['A']+B_n/ratio['B']+C_n/ratio['C']

fixed_loss = 0.95
capac=(prod[0]/2)*fixed_loss
for i in range(24):
    capac=min(capac,prod[i]*fixed_loss)



#################################################################################################################################################################



ratio_type={}
price_type={}

ratio_type['A']=0.6
ratio_type['B']=0.66
ratio_type['C']=0.72
price_type['A']=1.2
price_type['B']=1.1
price_type['C']=1

stock_for_cost=np.zeros(24)

status_list=[]

result_4_1=np.zeros((402,24))
min_4_1=[]

for index in range(24):
    costLP=pulp.LpProblem("cost_problem_"+str(index),sense=pulp.LpMinimize)

    yield_list=[]

    for i in range(402):
        if i < 10:
            yield_list.append(pulp.LpVariable('00'+str(i),lowBound=0,upBound=40000,cat='Integer'))
        elif i<100:
            yield_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=40000,cat='Integer'))
        else:
            yield_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=40000,cat='Integer'))

    costLP += pulp.lpSum([(yield_list[i]*price_type[data[i][0]])for i in range(402)])

    costLP += ((pulp.lpSum([(yield_list[i]*(ratio_type[data[i][0]])**(-1) )for i in range(402)])+stock_for_cost[index]-capac) >= capac)

    for i in range(402):
        costLP += (yield_list[i]<=Bound[i][index])

    costLP.solve()
    print(costLP.name)
    print("Status :", pulp.LpStatus[costLP.status])  # 输出求解状态
    count=0
    for v in costLP.variables():
        result_4_1[count][index]=v.varValue
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值
        count+=1

    print("Min f(x) =", pulp.value(costLP.objective))  # 输出最优解的目标函数值
    min_4_1.append(pulp.value(costLP.objective))
    status_list.append(pulp.LpStatus[costLP.status])

    sum_prod=0
    count=0
    for v in costLP.variables():
        sum_prod+=v.varValue/ratio_type[data[count][0]]
        count+=1
    if index <23:
        stock_for_cost[index+1]=stock_for_cost[index]+sum_prod-capac



print(status_list)
print(min_4_1)

##############################################################################################################################
    
dealt_shipper=np.zeros((8,24))

avg_list=[1,2,3]

for i in range(8):
    if i in avg_list:
        count=0
        Sum=0
        for k in range(240):
            if data_2[i][k]>0:
                count+=1
                Sum+=data_2[i][k]
        Sum=Sum/count
        for j in range(24):
            dealt_shipper[i][j]=Sum
    else:
        for j in range(24):
            count=0
            for k in range(10):
                if data_2[i][24*k+j]!=0:
                    dealt_shipper[i][j]+=data_2[i][24*k+j]
                    count+=1
            if count==0:
                c=0
                for k in range(240):
                    if data_2[i][k]!=0:
                        c+=1
                        dealt_shipper[i][j]+=data_2[i][k]
                dealt_shipper[i][j]/=c
            else:
                dealt_shipper[i][j]/=count

result_4_2=[]

status_4_2=[]

for index in range(24):

    lossLP=pulp.LpProblem("loss_problem_"+str(index),sense=pulp.LpMinimize)

    dec_list=[]

    par=np.zeros(402*8)

    for i in range(402*8):
        par[i]=(dealt_shipper[i%8][index]/100)/ratio_type[data[int(i/8)][0]]

    #print(dealt_shipper)
    #print(par)

    for i in range(8*402):
        if i < 10:
            dec_list.append(pulp.LpVariable('000'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        elif i <100:
            dec_list.append(pulp.LpVariable('00'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        elif i <1000:
            dec_list.append(pulp.LpVariable('0'+str(i),lowBound=0,upBound=6000,cat='Integer'))
        else:
            dec_list.append(pulp.LpVariable(str(i),lowBound=0,upBound=6000,cat='Integer'))
       
    for i in range(402):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for j in range(8)])==result_4_1[i][index])

    for j in range(8):
        lossLP += (pulp.lpSum([(dec_list[i*8+j])for i in range(402)])<=6000)
    
    lossLP += pulp.lpSum([(dec_list[i]*par[i])  for i in range(402*8)])

    
    lossLP.solve()
    
    print(lossLP.name)
    print("Status :", pulp.LpStatus[lossLP.status])  # 输出求解状态
    for v in lossLP.variables():
        print(v.name, "=", v.varValue)  # 输出每个变量的最优值

    print("Min f(x) =", pulp.value(lossLP.objective))  # 输出最优解的目标函数值
    
    rlt=np.zeros((402,8))

    #print(par)
    count=0
    for v in lossLP.variables():
        rlt[int(count/8)][count%8]=v.varValue
        count+=1
        #print(v.name, "=", v.varValue)  # 输出每个变量的最优值
    status_4_2.append(pulp.LpStatus[lossLP.status])
    
    result_4_2.append(rlt)
    #print(rlt)

print(status_4_2)

display_4_1=np.zeros((402,24))

for i in range(402):
    for j in range(24):
        display_4_1[i][j]=int(result_4_1[i][j]/prob[i])

print(display_4_1)
print(capac)

workbook=openpyxl.load_workbook("附件A 订购方案数据结果.xlsx")
worksheet=workbook.worksheets[2]

for i in range(402):
    for j in range(24):
        worksheet.cell(7+i,2+j,display_4_1[i][j])



workbook.save(filename="附件A 订购方案数据结果.xlsx")

workbook=openpyxl.load_workbook("附件B 转运方案数据结果.xlsx")
worksheet=workbook.worksheets[2]


for i in range(402):
    for j in range(24):
        for k in range(8):
            worksheet.cell(7+i,2+8*j+k,result_4_2[j][i][k])


workbook.save(filename="附件B 转运方案数据结果.xlsx")
